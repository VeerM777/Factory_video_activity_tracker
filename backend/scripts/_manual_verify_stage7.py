"""One-off manual check: build a workbook via Stage 7 from ground-truth
inputs and confirm it's structurally sound (run through recalc separately)."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl

from app.models.schemas import Classification, Segment
from app.pipeline.stage6_tmu_engine import build_most_row
from app.pipeline.stage7_excel_writer import write_most_analysis_workbook

ROOT = Path(__file__).parent.parent.parent
GROUND_TRUTH = ROOT / "data" / "ground_truth" / "ASSY_WITH_PRESS_OPERATION_ground_truth.xlsx"
TEMPLATE = ROOT / "data" / "templates" / "most_analysis_template.xlsx"
OUT = Path(
    r"C:\Users\PREMRA~1\AppData\Local\Temp\claude\C--Users-premraval010-Desktop-Factory-Video-Analysis\72ae13bb-373e-43e4-9204-b98d21e5f832\scratchpad\stage7_verify_output.xlsx"
)

wb = openpyxl.load_workbook(GROUND_TRUTH, data_only=True)
ws = wb["MOST Analysis"]

rows = []
param_cols = ["F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]
for r in range(6, ws.max_row + 1):
    if ws[f"A{r}"].value is None:
        continue
    data_card = ws[f"E{r}"].value
    values = [ws[f"{c}{r}"].value for c in param_cols]
    values = [v for v in values if v is not None]
    seg = Segment(
        segment_id=r,
        source_video_uri="gs://bucket/assy_press_op.mp4",
        t_start_sec=float(r),
        t_end_sec=float(r) + 1.0,
        description=ws[f"T{r}"].value,
        model_version="gemini-2.5-pro",
        prompt_version="v1",
    )
    cls = Classification(
        data_card=data_card,
        param_values=values,
        muda_ref=ws[f"V{r}"].value,
        freq=ws[f"R{r}"].value,
        confidence=0.95,
        model_version="gemini-2.5-pro",
        prompt_version="v1",
    )
    rows.append(build_most_row(seg, cls, s_no=r - 5, activity_description="ASSY WITH PRESS OPERATION"))

write_most_analysis_workbook(rows, TEMPLATE, OUT, activity_description="ASSY WITH PRESS OPERATION")
print(f"wrote {len(rows)} rows to {OUT}")
