"""Stage 7 -- Excel Generation.

Copies the real target template (never rebuilds it) and writes one row per
MostRow into 'MOST Analysis'. Formula columns (Q, S, W, AC) are written as
live Excel formulas -- translated per-row from the template's own formula
text -- so the workbook keeps recalculating if a reviewer edits a cell by
hand in Stage 8. Only the taxonomy-bucket columns (Y/Z/AA/AB) intentionally
use a classification-name lookup instead of the template's numeric-ref-range
IF()s -- see the note below.

Sheet 2 ('VA SVA NVA Summary') is left completely untouched: its SUMIF/pivot
formulas already reference whole columns on Sheet 1, so they keep working
unmodified as rows are added.

Two latent bugs in the original workbook's Y/Z/AA/AB formulas, found while
building this (neither is exercised by the example workbook's own 26 rows,
so they were invisible until now):
  1. Ref 0 (Noise) satisfies "V<23" and is silently counted as NVA time.
  2. Ref 23 ("Manual testing", taxonomy classification NVA) satisfies
     none of Y/Z/AA/AB's range conditions and is silently dropped
     from every bucket.
Both stem from bucketing by numeric ref ranges instead of the taxonomy's own
classification field. Since the taxonomy is required to be versioned/editable
(refs can move -- ref 48/50 already did in this build), this writer buckets
new rows via VLOOKUP against the taxonomy's classification column instead,
which is immune to both bugs and to any future renumbering.
"""
from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

from app.config.most_tables import load_most_tables
from app.models.schemas import MostRow

SHEET_NAME = "MOST Analysis"
FIRST_DATA_ROW = 6
MAX_CLEAR_COL = 45  # generous headroom above AM (col 39)

# Extra traceability & activity columns, appended after the template's own AC (col 29).
TRACE_HEADERS = {
    "AD": "Source Video",
    "AE": "Segment Start (s)",
    "AF": "Segment End (s)",
    "AG": "Segmentation Model",
    "AH": "Classification Model",
    "AI": "Confidence",
    "AJ": "Human Corrected",
    "AK": "Activity & Movement Details",
    "AL": "Activity Duration (sec)",
    "AM": "Activity Timeline",
    "AN": "Elemental Description",
}


def _clear_existing_data_rows(ws) -> None:
    from openpyxl.utils import get_column_letter

    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        for c in range(1, MAX_CLEAR_COL + 1):
            ws[f"{get_column_letter(c)}{r}"] = None


def _write_trace_headers(ws) -> None:
    for col, header in TRACE_HEADERS.items():
        ws[f"{col}5"] = header


def _create_timeline_chart_sheet(wb, rows: list[MostRow]) -> None:
    """Creates a dedicated 'Activity Timeline Chart' worksheet with a visual Gantt chart
    displaying the chronological distribution, durations, and category breakdowns of all activities."""
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    chart_sheet_name = "Activity Timeline Chart"
    if chart_sheet_name in wb.sheetnames:
        del wb[chart_sheet_name]

    ws_chart = wb.create_sheet(title=chart_sheet_name)
    ws_chart.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_chart.merge_cells("A1:G1")
    title_cell = ws_chart["A1"]
    title_cell.value = "Activity Timeline & Distribution Chart"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_chart.row_dimensions[1].height = 35

    # Headers
    headers = [
        "Activity Description",
        "Start Time (s)",
        "Duration (s)",
        "End Time (s)",
        "Category",
        "Movement State",
        "Machine State",
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws_chart.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_chart.row_dimensions[3].height = 25

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    start_row = 4
    for i, r in enumerate(rows):
        row_num = start_row + i
        ws_chart.cell(row=row_num, column=1, value=r.elemental_description)
        ws_chart.cell(row=row_num, column=2, value=r.t_start_sec)
        ws_chart.cell(row=row_num, column=3, value=r.activity_duration_sec)
        ws_chart.cell(row=row_num, column=4, value=r.t_end_sec)
        ws_chart.cell(row=row_num, column=5, value=r.category)

        # Parse movement state & machine state from details string
        mov_state = "MOVE"
        mac_state = "IDLE"
        if "(" in r.activity_movement_details and ")" in r.activity_movement_details:
            mov_state = r.activity_movement_details.split("(")[1].split(")")[0]
        if "Machine:" in r.activity_movement_details:
            mac_state = r.activity_movement_details.split("Machine:")[1].strip()

        ws_chart.cell(row=row_num, column=6, value=mov_state)
        ws_chart.cell(row=row_num, column=7, value=mac_state)

        for col in range(1, 8):
            c = ws_chart.cell(row=row_num, column=col)
            c.border = thin_border
            if col in (2, 3, 4):
                c.alignment = Alignment(horizontal="right")
            else:
                c.alignment = Alignment(horizontal="left")

    num_rows = len(rows)
    end_row = start_row + num_rows - 1

    # Create Horizontal Stacked Bar Chart for Gantt Timeline representation
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = "Activity Execution Timeline (Gantt Chart)"
    chart.height = max(10, num_rows * 0.8)
    chart.width = 22

    # Series: Start Time (invisible offset) and Duration (active bar)
    data = Reference(ws_chart, min_col=2, min_row=3, max_col=3, max_row=end_row)
    cats = Reference(ws_chart, min_col=1, min_row=4, max_row=end_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    chart.x_axis.title = "Timeline (Seconds)"
    chart.y_axis.title = "Activities"
    chart.legend = None

    # Embed chart beside table
    ws_chart.add_chart(chart, "I3")

    # Column widths
    ws_chart.column_dimensions["A"].width = 38
    ws_chart.column_dimensions["B"].width = 15
    ws_chart.column_dimensions["C"].width = 15
    ws_chart.column_dimensions["D"].width = 15
    ws_chart.column_dimensions["E"].width = 25
    ws_chart.column_dimensions["F"].width = 18
    ws_chart.column_dimensions["G"].width = 18


def write_most_analysis_workbook(
    rows: list[MostRow],
    template_path: Path,
    output_path: Path,
    activity_description: str,
) -> Path:
    if not rows:
        raise ValueError("no rows to write")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        shutil.copyfile(template_path, output_path)
    except PermissionError:
        import time
        output_path = output_path.with_name(f"{output_path.stem}_{int(time.time())}{output_path.suffix}")
        shutil.copyfile(template_path, output_path)

    wb = load_workbook(output_path)
    ws = wb[SHEET_NAME]

    _clear_existing_data_rows(ws)
    _write_trace_headers(ws)

    tables = load_most_tables()
    q_template = "=IF(E6=$E$1,($F$1&F6&$G$1&G6&$H$1&H6&$I$1&I6&$J$1&J6&$K$1&K6&$L$1&L6),IF(E6=$E$2,($F$2&F6&$G$2&G6&$H$2&H6&$I$2&I6&$J$2&J6&$K$2&K6&$L$2&L6),IF(E6=$E$3,($F$3&F6&$G$3&G6&$H$3&H6&$I$3&I6&$J$3&J6&$K$3&K6&$L$3&L6&$M$3&M6&$N$3&N6&$O$3&O6&$P$3&P6),IF(E6=$E$4,F6&\"SEC\"))))"
    s_template = "=IF(E6=$E$4,R6*F6/0.036,(SUM(F6:P6)*10*R6))"
    ac_template = "=VLOOKUP(V6,'VA SVA NVA Summary'!A:E,3,0)"

    for i, row in enumerate(rows):
        r = FIRST_DATA_ROW + i
        model = tables.sequence_models[row.data_card]

        ws[f"A{r}"] = row.s_no
        ws[f"B{r}"] = row.station_no
        ws[f"C{r}"] = row.activity_no
        ws[f"D{r}"] = activity_description
        ws[f"E{r}"] = row.data_card
        for col, value in zip(model.columns, row.param_values):
            ws[f"{col}{r}"] = value

        ws[f"Q{r}"] = Translator(q_template, origin="Q6").translate_formula(f"Q{r}")
        ws[f"R{r}"] = row.freq
        ws[f"S{r}"] = Translator(s_template, origin="S6").translate_formula(f"S{r}")
        ws[f"T{r}"] = row.elemental_description
        ws[f"U{r}"] = row.operator
        ws[f"V{r}"] = row.muda_ref
        ws[f"W{r}"] = f"=S{r}*0.036"
        ws[f"X{r}"] = row.online_offline_mode

        classification_lookup = f"VLOOKUP(V{r},'VA SVA NVA Summary'!$A:$E,2,0)"
        ws[f"Y{r}"] = f'=IF({classification_lookup}="VA",W{r},0)'
        ws[f"Z{r}"] = f'=IF({classification_lookup}="NVA-N",W{r},0)'
        ws[f"AA{r}"] = f'=IF({classification_lookup}="SVA",W{r},0)'
        ws[f"AB{r}"] = f'=IF({classification_lookup}="NVA",W{r},0)'
        ws[f"AC{r}"] = Translator(ac_template, origin="AC6").translate_formula(f"AC{r}")

        ws[f"AD{r}"] = row.source_video_uri
        ws[f"AE{r}"] = row.t_start_sec
        ws[f"AF{r}"] = row.t_end_sec
        ws[f"AG{r}"] = row.segment_model_version
        ws[f"AH{r}"] = row.classification_model_version
        ws[f"AI{r}"] = row.confidence
        ws[f"AJ{r}"] = "YES" if row.human_corrected else "NO"
        ws[f"AK{r}"] = row.activity_movement_details
        ws[f"AL{r}"] = row.activity_duration_sec
        ws[f"AM{r}"] = row.activity_timeline
        ws[f"AN{r}"] = row.uppercase_elemental_description

    last_row = FIRST_DATA_ROW + len(rows) - 1
    ws["D6"] = activity_description
    ws["W4"] = f"=SUM(W{FIRST_DATA_ROW}:W{last_row})"

    # Generate dedicated Activity Timeline Chart worksheet tab
    _create_timeline_chart_sheet(wb, rows)

    wb.save(output_path)
    return output_path
