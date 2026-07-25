"""Phase 0 Definition of Done: an automated row-by-row diff report comparing
generated pipeline output against the hand-built ground-truth workbook.

Rows are compared by position (generated row i vs. ground-truth row i) --
correct as long as segmentation produces the same number of motions in the
same order as the human-built study, which is exactly what Phase 0 is
supposed to prove or disprove.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from pydantic import BaseModel

from app.models.schemas import MostRow


class GroundTruthRow(BaseModel):
    row_index: int
    data_card: str
    param_values: list[int]
    freq: int
    muda_ref: int
    tmu: float
    total_time_sec: float
    description: str


class RowDiff(BaseModel):
    row_index: int
    generated_tmu: float | None
    ground_truth_tmu: float | None
    tmu_delta: float | None
    generated_muda_ref: int | None
    ground_truth_muda_ref: int | None
    category_match: bool
    generated_description: str | None
    ground_truth_description: str | None


class DiffReport(BaseModel):
    generated_row_count: int
    ground_truth_row_count: int
    segment_count_match: bool
    row_diffs: list[RowDiff]
    total_tmu_delta: float
    category_mismatch_count: int

    def summary_text(self) -> str:
        lines = [
            f"Segment count: generated={self.generated_row_count} "
            f"ground_truth={self.ground_truth_row_count} "
            f"({'MATCH' if self.segment_count_match else 'MISMATCH'})",
            f"Total TMU delta across matched rows: {self.total_tmu_delta:.3f}",
            f"Category mismatches: {self.category_mismatch_count}/{len(self.row_diffs)}",
            "",
        ]
        for d in self.row_diffs:
            flag = "OK" if d.category_match and (d.tmu_delta is None or abs(d.tmu_delta) < 1e-6) else "DIFF"
            lines.append(
                f"[{flag}] row {d.row_index}: tmu {d.generated_tmu} vs {d.ground_truth_tmu} "
                f"(delta {d.tmu_delta}); ref {d.generated_muda_ref} vs {d.ground_truth_muda_ref}"
            )
        return "\n".join(lines)


def _load_ground_truth_rows(ground_truth_xlsx: Path) -> list[GroundTruthRow]:
    wb = openpyxl.load_workbook(ground_truth_xlsx, data_only=True)
    ws = wb["MOST Analysis"]
    param_cols = ["F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]
    rows = []
    for r in range(6, ws.max_row + 1):
        if ws[f"A{r}"].value is None:
            continue
        values = [ws[f"{c}{r}"].value for c in param_cols]
        values = [v for v in values if v is not None]
        rows.append(
            GroundTruthRow(
                row_index=r - 6,
                data_card=ws[f"E{r}"].value,
                param_values=values,
                freq=ws[f"R{r}"].value,
                muda_ref=ws[f"V{r}"].value,
                tmu=ws[f"S{r}"].value,
                total_time_sec=ws[f"W{r}"].value,
                description=ws[f"T{r}"].value,
            )
        )
    return rows


def build_diff_report(generated_rows: list[MostRow], ground_truth_xlsx: Path) -> DiffReport:
    gt_rows = _load_ground_truth_rows(ground_truth_xlsx)

    row_diffs: list[RowDiff] = []
    total_delta = 0.0
    mismatches = 0

    max_len = max(len(generated_rows), len(gt_rows))
    for i in range(max_len):
        gen = generated_rows[i] if i < len(generated_rows) else None
        gt = gt_rows[i] if i < len(gt_rows) else None

        tmu_delta = (gen.tmu - gt.tmu) if (gen and gt) else None
        category_match = (gen.muda_ref == gt.muda_ref) if (gen and gt) else False
        if not category_match:
            mismatches += 1
        if tmu_delta is not None:
            total_delta += abs(tmu_delta)

        row_diffs.append(
            RowDiff(
                row_index=i,
                generated_tmu=gen.tmu if gen else None,
                ground_truth_tmu=gt.tmu if gt else None,
                tmu_delta=tmu_delta,
                generated_muda_ref=gen.muda_ref if gen else None,
                ground_truth_muda_ref=gt.muda_ref if gt else None,
                category_match=category_match,
                generated_description=gen.elemental_description if gen else None,
                ground_truth_description=gt.description if gt else None,
            )
        )

    return DiffReport(
        generated_row_count=len(generated_rows),
        ground_truth_row_count=len(gt_rows),
        segment_count_match=len(generated_rows) == len(gt_rows),
        row_diffs=row_diffs,
        total_tmu_delta=total_delta,
        category_mismatch_count=mismatches,
    )
