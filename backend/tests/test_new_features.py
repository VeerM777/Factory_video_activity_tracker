"""Unit tests for Activity-Change details, 3 new Excel columns, Camera Calibration,
Human Review Stage 8, Feedback Loop Stage 9, and FastAPI application endpoints.
"""
from __future__ import annotations

import tempfile
from pathlib import Path
from openpyxl import load_workbook
import pytest

from app.models.schemas import Classification, Segment, MostRow, ReviewFlag
from app.pipeline.stage6_tmu_engine import build_most_row
from app.pipeline.stage7_excel_writer import write_most_analysis_workbook
from app.services.camera_calibration import PixelToMetricMapper
from app.pipeline.stage8_human_review import HumanReviewEngine
from app.pipeline.stage9_feedback_loop import FeedbackLoopEngine, FeedbackExemplar


def test_activity_details_duration_and_timeline():
    segment = Segment(
        segment_id=0,
        source_video_uri="test.mp4",
        t_start_sec=4.20,
        t_end_sec=6.50,
        description="GRASP GLUING TOOL",
        human_movement_state="GRASP",
        machine_state="IDLE",
        model_version="test",
        prompt_version="test",
    )
    classification = Classification(
        data_card="G",
        param_values=[1, 0, 1, 1, 0, 1, 0],
        muda_ref=35,
        freq=1,
        confidence=0.95,
        model_version="test",
        prompt_version="test",
    )
    row = build_most_row(segment, classification, s_no=1, activity_description="GLUING OP")

    assert row.activity_duration_sec == 2.30
    assert row.activity_timeline == "4.20s - 6.50s"
    assert "Human: GRASP GLUING TOOL (GRASP) | Machine: IDLE" in row.activity_movement_details


def test_excel_writer_new_columns():
    segment = Segment(
        segment_id=0,
        source_video_uri="test.mp4",
        t_start_sec=1.0,
        t_end_sec=3.5,
        description="PLACE WORKPIECE",
        human_movement_state="MOVE",
        machine_state="IDLE",
        model_version="test",
        prompt_version="test",
    )
    classification = Classification(
        data_card="G",
        param_values=[1, 0, 1, 1, 0, 1, 0],
        muda_ref=1,
        freq=1,
        confidence=0.9,
        model_version="test",
        prompt_version="test",
    )
    row = build_most_row(segment, classification, s_no=1, activity_description="TEST")

    root = Path(__file__).parent.parent.parent
    template = root / "data" / "templates" / "most_analysis_template.xlsx"

    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = Path(tmpdir) / "output.xlsx"
        write_most_analysis_workbook([row], template, out_path, "TEST")

        wb = load_workbook(out_path)
        ws = wb["MOST Analysis"]

        # Header checks
        assert ws["AK5"].value == "Activity & Movement Details"
        assert ws["AL5"].value == "Activity Duration (sec)"
        assert ws["AM5"].value == "Activity Timeline"

        # Row 6 data checks
        assert "Human: PLACE WORKPIECE (MOVE) | Machine: IDLE" in str(ws["AK6"].value)
        assert ws["AL6"].value == 2.5
        assert ws["AM6"].value == "1.00s - 3.50s"


def test_camera_calibration():
    # 100 pixels = 10 cm => 10 px/cm
    mapper = PixelToMetricMapper.from_reference_object(pixel_length=100.0, real_length_cm=10.0)
    assert mapper.px_to_cm(50.0) == 5.0
    assert mapper.px_to_m(500.0) == 0.5


def test_human_review_engine():
    segment = Segment(
        segment_id=0,
        source_video_uri="test.mp4",
        t_start_sec=0.0,
        t_end_sec=2.0,
        description="REACH TOOL",
        model_version="v1",
        prompt_version="v1",
    )
    flag = ReviewFlag(segment_id=0, reason="Low confidence", confidence=0.5)

    engine = HumanReviewEngine(rows=[], segments=[segment], review_flags=[flag])
    assert len(engine.get_pending_flags()) == 1

    # Human corrects classification
    updated_row = engine.update_row_classification(
        segment_id=0,
        data_card="G",
        param_values=[1, 0, 1, 1, 0, 1, 0],
        muda_ref=1,
    )

    assert updated_row.human_corrected is True
    assert engine.resolve_all_clear() is True
    assert len(engine.get_finalized_rows()) == 1


def test_feedback_loop_engine():
    with tempfile.TemporaryDirectory() as tmpdir:
        lib_path = Path(tmpdir) / "feedback.json"
        engine = FeedbackLoopEngine(library_path=lib_path)

        segment = Segment(
            segment_id=0,
            source_video_uri="test.mp4",
            t_start_sec=0.0,
            t_end_sec=2.0,
            description="GRASP GLUING TOOL",
            model_version="v1",
            prompt_version="v1",
        )
        classification = Classification(
            data_card="G",
            param_values=[1, 0, 1, 1, 0, 1, 0],
            muda_ref=35,
            confidence=1.0,
            model_version="v1",
            prompt_version="v1",
        )
        row = build_most_row(segment, classification, s_no=1, activity_description="GLUE")
        row.human_corrected = True

        engine.record_correction(row)
        assert len(engine.exemplars) == 1

        prompt_ctx = engine.format_few_shot_prompt_context()
        assert "HUMAN-VERIFIED FEW-SHOT EXAMPLES" in prompt_ctx
        assert "GRASP GLUING TOOL" in prompt_ctx
