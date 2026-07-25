"""Validates the Stage 6 engine against the real ground-truth workbook.

Loads ASSY WITH PRESS OPERATION.xlsx's cached (already-computed) values
directly -- no re-derivation -- and asserts our pure-code engine reproduces
TMU, Total Time, and the VA/SVA/NVA/NVA-N bucket for every row, plus the
sheet-wide grand total. This is the Phase 0 "row-by-row diff" proof that the
deterministic engine matches the hand-built spreadsheet.
"""
from pathlib import Path

import openpyxl
import pytest

from app.models.schemas import Classification, Segment
from app.pipeline.stage6_tmu_engine import build_most_row, compute_tmu

GROUND_TRUTH = (
    Path(__file__).parent.parent.parent
    / "data"
    / "ground_truth"
    / "ASSY_WITH_PRESS_OPERATION_ground_truth.xlsx"
)


def _load_ground_truth_rows() -> list[dict]:
    wb = openpyxl.load_workbook(GROUND_TRUTH, data_only=True)
    ws = wb["MOST Analysis"]
    rows = []
    for r in range(6, ws.max_row + 1):
        if ws[f"A{r}"].value is None:
            continue
        param_cols = ["F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]
        param_values = [ws[f"{c}{r}"].value for c in param_cols]
        param_values = [v for v in param_values if v is not None]
        rows.append(
            {
                "row": r,
                "data_card": ws[f"E{r}"].value,
                "param_values": param_values,
                "freq": ws[f"R{r}"].value,
                "muda_ref": ws[f"V{r}"].value,
                "expected_tmu": ws[f"S{r}"].value,
                "expected_total_time_sec": ws[f"W{r}"].value,
                "expected_va": ws[f"Y{r}"].value,
                "expected_nvan": ws[f"Z{r}"].value,
                "expected_sva": ws[f"AA{r}"].value,
                "expected_nva": ws[f"AB{r}"].value,
                "expected_grand_total_sec": ws["W4"].value,
            }
        )
    return rows


GROUND_TRUTH_ROWS = _load_ground_truth_rows()


@pytest.mark.parametrize("row", GROUND_TRUTH_ROWS, ids=lambda r: f"row{r['row']}")
def test_tmu_matches_ground_truth(row):
    tmu = compute_tmu(row["data_card"], row["param_values"], row["freq"])
    assert tmu == pytest.approx(row["expected_tmu"])


@pytest.mark.parametrize("row", GROUND_TRUTH_ROWS, ids=lambda r: f"row{r['row']}")
def test_full_row_matches_ground_truth(row):
    segment = Segment(
        segment_id=row["row"],
        source_video_uri="test://ground-truth",
        t_start_sec=0.0,
        t_end_sec=1.0,
        description="test segment",
        model_version="test",
        prompt_version="test",
    )
    classification = Classification(
        data_card=row["data_card"],
        param_values=row["param_values"],
        muda_ref=row["muda_ref"],
        freq=row["freq"],
        confidence=1.0,
        model_version="test",
        prompt_version="test",
    )
    most_row = build_most_row(
        segment, classification, s_no=row["row"], activity_description="ASSY WITH PRESS OPERATION"
    )

    assert most_row.tmu == pytest.approx(row["expected_tmu"])
    assert most_row.total_time_sec == pytest.approx(row["expected_total_time_sec"])

    # Ground truth's own bucketing has the ref-range bug described in
    # stage6_tmu_engine's docstring, but none of this workbook's real rows
    # (all NVA/SVA/VA, no Noise) fall in the affected range, so bucketed
    # amounts should still match exactly.
    assert most_row.va_sec == pytest.approx(row["expected_va"])
    assert most_row.nvan_sec == pytest.approx(row["expected_nvan"])
    assert most_row.sva_sec == pytest.approx(row["expected_sva"])
    assert most_row.nva_sec == pytest.approx(row["expected_nva"])


def test_grand_total_matches_ground_truth():
    total = sum(
        compute_tmu(r["data_card"], r["param_values"], r["freq"]) * 0.036 for r in GROUND_TRUTH_ROWS
    )
    assert total == pytest.approx(GROUND_TRUTH_ROWS[0]["expected_grand_total_sec"])


def test_process_time_formula():
    # No PT rows exist in the ground-truth sample; verify the documented
    # formula (FREQ * seconds / 0.036) directly.
    tmu = compute_tmu("PT", [12], freq=2)
    assert tmu == pytest.approx(2 * 12 / 0.036)


def test_rejects_out_of_table_index_value():
    classification = Classification(
        data_card="G",
        param_values=[999, 0, 1, 0, 0, 0, 0],
        muda_ref=5,
        freq=1,
        confidence=0.9,
        model_version="test",
        prompt_version="test",
    )
    with pytest.raises(ValueError):
        classification.validate_against_tables()


def test_rejects_unknown_taxonomy_ref():
    classification = Classification(
        data_card="G",
        param_values=[1, 0, 1, 0, 0, 0, 0],
        muda_ref=9999,
        freq=1,
        confidence=0.9,
        model_version="test",
        prompt_version="test",
    )
    with pytest.raises(ValueError):
        classification.validate_against_tables()
