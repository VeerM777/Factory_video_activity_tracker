"""Sanity check: diffing the ground truth against itself must show a perfect
match (0 TMU delta, 0 category mismatches, segment counts equal)."""
from pathlib import Path

from app.models.schemas import Classification, Segment
from app.pipeline.diff_report import build_diff_report
from app.pipeline.stage6_tmu_engine import build_most_row

GROUND_TRUTH = (
    Path(__file__).parent.parent.parent
    / "data"
    / "ground_truth"
    / "ASSY_WITH_PRESS_OPERATION_ground_truth.xlsx"
)


def _rows_from_ground_truth() -> list:
    import openpyxl

    wb = openpyxl.load_workbook(GROUND_TRUTH, data_only=True)
    ws = wb["MOST Analysis"]
    param_cols = ["F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]
    rows = []
    for r in range(6, ws.max_row + 1):
        if ws[f"A{r}"].value is None:
            continue
        values = [ws[f"{c}{r}"].value for c in param_cols]
        values = [v for v in values if v is not None]
        seg = Segment(
            segment_id=r,
            source_video_uri="test://ground-truth",
            t_start_sec=0.0,
            t_end_sec=1.0,
            description=ws[f"T{r}"].value,
            model_version="test",
            prompt_version="test",
        )
        cls = Classification(
            data_card=ws[f"E{r}"].value,
            param_values=values,
            muda_ref=ws[f"V{r}"].value,
            freq=ws[f"R{r}"].value,
            confidence=1.0,
            model_version="test",
            prompt_version="test",
        )
        rows.append(build_most_row(seg, cls, s_no=r - 5, activity_description="ASSY WITH PRESS OPERATION"))
    return rows


def test_self_diff_is_perfect_match():
    rows = _rows_from_ground_truth()
    report = build_diff_report(rows, GROUND_TRUTH)

    assert report.segment_count_match
    assert report.generated_row_count == report.ground_truth_row_count == 26
    assert report.category_mismatch_count == 0
    assert report.total_tmu_delta == 0.0
    assert all(d.category_match for d in report.row_diffs)
